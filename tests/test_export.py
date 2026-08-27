"""
Test su build_final_table: merge corretto di metadati, inventario,
storico e forecast con prefisso 'f' sulle colonne previste.
"""

import pandas as pd

from forecast_lib.export import build_final_table, build_forecast_wide


def test_build_forecast_wide_orders_columns():
    df_long = pd.DataFrame({
        "SKU": ["A", "A", "A"],
        "Period": ["2025_03", "2025_01", "2025_02"],
        "Forecast": [30, 10, 20],
    })
    out = build_forecast_wide(df_long, id_col="SKU")
    # le colonne forecast devono essere ordinate cronologicamente
    cols = [c for c in out.columns if c != "SKU"]
    assert cols == ["2025_01", "2025_02", "2025_03"]


def test_build_final_table_basic_merge():
    # storico
    df_filtered = pd.DataFrame({
        "SKU": ["A", "A", "B", "B"],
        "Description": ["aaa", "aaa", "bbb", "bbb"],
        "Round": [6, 6, 12, 12],
        "BUn": ["EA", "EA", "EA", "EA"],
        "Period": ["2024_01", "2024_02"] * 2,
        "Demand": [10, 15, 20, 25],
        "Date": pd.to_datetime(["2024_01", "2024_02"] * 2, format="%Y_%m"),
    })

    # forecast wide
    df_fc_wide = pd.DataFrame({
        "SKU": ["A", "B"],
        "2025_01": [12.0, 22.0],
        "2025_02": [14.0, 24.0],
    })

    # inventory
    df_inventory = pd.DataFrame({
        "SKU": ["A", "B"],
        "LT_Final": [30, 45],
        "ABC": ["A", "B"],
        "XYZ": ["X", "Y"],
        "SafetyStock": [12.0, 0.0],
    })

    out = build_final_table(
        df_filtered, df_fc_wide, df_inventory,
        id_col="SKU", desc_col="Description",
        pack_size_col="Round", uom_col="BUn",
    )

    # SKU come stringa (dtype puo' essere object o StringDtype in pandas moderno)
    assert all(isinstance(v, str) for v in out["SKU"])

    # colonne forecast devono avere prefisso 'f'
    assert "f2025_01" in out.columns
    assert "f2025_02" in out.columns

    # colonne storiche senza prefisso
    assert "2024_01" in out.columns
    assert "2024_02" in out.columns

    # LT_Final rinominato in LT
    assert "LT" in out.columns

    # 2 righe (una per SKU)
    assert len(out) == 2


def test_build_final_table_without_inventory():
    df_filtered = pd.DataFrame({
        "SKU": ["A"],
        "Description": ["aaa"],
        "Round": [6],
        "BUn": ["EA"],
        "Period": ["2024_01"],
        "Demand": [10],
        "Date": pd.to_datetime(["2024_01"], format="%Y_%m"),
    })
    df_fc_wide = pd.DataFrame({"SKU": ["A"], "2025_01": [12.0]})

    out = build_final_table(
        df_filtered, df_fc_wide, df_inventory=None,
        id_col="SKU", desc_col="Description",
        pack_size_col="Round", uom_col="BUn",
    )

    # niente colonne inventario
    assert "LT" not in out.columns
    assert "ABC" not in out.columns
    assert "SafetyStock" not in out.columns
    # ma forecast e storico ci sono
    assert "f2025_01" in out.columns
    assert "2024_01" in out.columns


def test_build_final_table_fills_missing_safety_stock_with_zero():
    # SKU 'B' non e' nell'inventario -> SafetyStock NaN -> deve diventare 0
    df_filtered = pd.DataFrame({
        "SKU": ["A", "B"],
        "Description": ["aaa", "bbb"],
        "Round": [6, 12],
        "BUn": ["EA", "EA"],
        "Period": ["2024_01", "2024_01"],
        "Demand": [10, 20],
        "Date": pd.to_datetime(["2024_01", "2024_01"], format="%Y_%m"),
    })
    df_fc_wide = pd.DataFrame({"SKU": ["A", "B"], "2025_01": [12.0, 22.0]})
    df_inventory = pd.DataFrame({
        "SKU": ["A"],   # solo A
        "LT_Final": [30],
        "ABC": ["A"], "XYZ": ["X"], "SafetyStock": [10.0],
    })

    out = build_final_table(
        df_filtered, df_fc_wide, df_inventory,
        id_col="SKU", desc_col="Description",
        pack_size_col="Round", uom_col="BUn",
    )

    b_row = out[out["SKU"] == "B"].iloc[0]
    assert b_row["SafetyStock"] == 0.0


# ======================================================================
# save_excel / build_run_info — foglio "Run info" (Fase 1.6)
# ======================================================================

import openpyxl

from forecast_lib.export import (
    DATA_SHEET_NAME,
    RUN_INFO_SHEET_NAME,
    build_run_info,
    run_info_to_frame,
    save_excel,
)


def _small_frame():
    return pd.DataFrame({"SKU": ["A", "B"], "f2026_01": [10.0, 20.0]})


def test_save_excel_without_run_info_writes_a_single_sheet(tmp_path):
    path = tmp_path / "out.xlsx"
    save_excel(_small_frame(), str(path))

    book = openpyxl.load_workbook(path)
    assert len(book.sheetnames) == 1
    assert RUN_INFO_SHEET_NAME not in book.sheetnames


def test_save_excel_with_run_info_writes_two_sheets_data_first(tmp_path):
    """Vincolo duro: la tabella dati resta il PRIMO foglio — questo stesso
    progetto legge l'input con list(all_sheets.keys())[0]."""
    path = tmp_path / "out.xlsx"
    save_excel(_small_frame(), str(path), run_info={"Campo di prova": "valore"})

    book = openpyxl.load_workbook(path)
    assert book.sheetnames == [DATA_SHEET_NAME, RUN_INFO_SHEET_NAME]


def test_save_excel_first_sheet_is_readable_without_sheet_name(tmp_path):
    """`pd.read_excel(path)` senza sheet_name legge il primo foglio: deve
    restituire i dati, non i metadati del run."""
    path = tmp_path / "out.xlsx"
    df = _small_frame()
    save_excel(df, str(path), run_info={"Campo di prova": "valore"})

    reread = pd.read_excel(path)
    assert list(reread.columns) == list(df.columns)
    assert reread["SKU"].tolist() == ["A", "B"]

    # e il consumatore che legge tutti i fogli trova i dati al primo posto
    all_sheets = pd.read_excel(path, sheet_name=None)
    assert list(all_sheets.keys())[0] == DATA_SHEET_NAME


def test_save_excel_run_info_sheet_contains_the_fields(tmp_path):
    path = tmp_path / "out.xlsx"
    save_excel(_small_frame(), str(path),
               run_info={"forecast_lib": "1.6.0", "INFERENCE_BATCH_SIZE": 32})

    info = pd.read_excel(path, sheet_name=RUN_INFO_SHEET_NAME)
    assert list(info.columns) == ["Campo", "Valore"]
    values = dict(zip(info["Campo"], info["Valore"]))
    assert values["forecast_lib"] == "1.6.0"
    assert values["INFERENCE_BATCH_SIZE"] == 32


def test_run_info_to_frame_renders_booleans_and_none():
    frame = run_info_to_frame({"Vero": True, "Falso": False, "Vuoto": None,
                               "Numero": 3.5, "Testo": "x"})
    values = dict(zip(frame["Campo"], frame["Valore"]))
    assert values["Vero"] == "si"
    assert values["Falso"] == "no"
    assert values["Vuoto"] == ""
    assert values["Numero"] == 3.5
    assert values["Testo"] == "x"


def test_run_info_to_frame_preserves_field_order():
    fields = {"primo": 1, "secondo": 2, "terzo": 3}
    assert run_info_to_frame(fields)["Campo"].tolist() == list(fields)


class _FakeModel:
    """Solo gli attributi `fl_*` che build_run_info legge dal modello."""
    fl_timesfm_tag = "v2.0.2"
    fl_pin_verified = False
    fl_model_revision = "1d95242"
    fl_device = "cuda (RTX 2070 SUPER)"
    fl_batch_size = 8
    fl_degraded = True
    fl_degraded_after_inference = True
    fl_inference_seconds = 12.5
    global_batch_size = 8


def test_build_run_info_reads_the_run_state_from_the_model():
    info = build_run_info(
        model=_FakeModel(),
        timestamp="2026-08-27 10:00:00",
        lib_version="1.6.0",
        timesfm_version="2.0.2",
        model_id="google/timesfm-2.5-200m-pytorch",
        model_revision="1d952420fba87f3c6dee4f240de0f1a0fbc790e3",
        inference_batch_size=32,
        q_global=0.57,
        kpi_motul=0.6978,
    )
    assert info["Pin TimesFM verificato"] is False
    assert info["TimesFM (tag pinnato)"] == "v2.0.2"
    assert info["Revision pesi richiesta"] == "1d952420fba87f3c6dee4f240de0f1a0fbc790e3"
    assert info["Revision pesi"] == "1d95242"
    assert info["Device"] == "cuda (RTX 2070 SUPER)"
    assert info["INFERENCE_BATCH_SIZE"] == 32
    assert info["Batch size effettivo"] == 8
    assert info["Degrado dopo inferenza reale"] is True
    assert info["Tempo di inferenza (s)"] == 12.5
    assert info["q globale (mediana shrinkage)"] == 0.57
    assert info["KPI Motul pesato"] == 0.6978


def test_build_run_info_without_a_model_leaves_the_fields_empty():
    """`RUN_BACKTEST = False` e i test chiamano build_run_info senza modello:
    non deve sollevare, e i campi del modello restano vuoti."""
    info = build_run_info(model=None, timestamp="2026-08-27 10:00:00",
                          timesfm_version="2.0.2")
    assert info["Pin TimesFM verificato"] is None
    assert info["Device"] is None
    # il fallback su timesfm_version copre il caso "modello non caricato"
    assert info["TimesFM (tag pinnato)"] == "2.0.2"


def test_build_run_info_is_writable_as_a_sheet(tmp_path):
    """Il dict deve attraversare run_info_to_frame senza tipi non scrivibili."""
    info = build_run_info(model=_FakeModel(), timestamp="2026-08-27 10:00:00",
                          lib_version="1.6.0", run_backtest=True,
                          n_backtest_skus=540, n_skus_zero_accuracy=31)
    path = tmp_path / "out.xlsx"
    save_excel(_small_frame(), str(path), run_info=info)

    sheet = pd.read_excel(path, sheet_name=RUN_INFO_SHEET_NAME)
    assert len(sheet) == len(info)
    values = dict(zip(sheet["Campo"], sheet["Valore"]))
    assert values["RUN_BACKTEST"] == "si"
    assert values["SKU con risultato di backtest"] == 540


def test_build_run_info_keeps_the_requested_revision_when_resolution_fails():
    """`_resolve_model_revision` e' diagnostica e non bloccante: torna None se si
    e' offline. Senza il campo "richiesta", il file non conserverebbe alcuna
    traccia di quali pesi erano stati chiesti — il dato su cui poggia la
    riproducibilita' del run."""
    class _NoRevision(_FakeModel):
        fl_model_revision = None

    info = build_run_info(model=_NoRevision(), model_revision="1d952420fba")

    assert info["Revision pesi"] is None
    assert info["Revision pesi richiesta"] == "1d952420fba"
